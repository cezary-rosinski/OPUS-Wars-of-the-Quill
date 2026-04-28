#!/usr/bin/env python3
"""
OPUS → EMLO + RDF converter
Usage: python converter.py OPUS_data_model.xlsx
"""

import re
import sys
import pandas as pd
from rdflib import Graph, Literal, Namespace, URIRef, BNode, RDF, RDFS, OWL, XSD
from rdflib.namespace import DCTERMS, SKOS
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# namespaces
OPUS   = Namespace("https://opus-project.eu/ontology/")
SCHEMA = Namespace("https://schema.org/")
BIBO   = Namespace("http://purl.org/ontology/bibo/")
GEO    = Namespace("http://www.w3.org/2003/01/geo/wgs84_pos#")

BASE = {
    "letter": "https://opus-project.eu/letter/",
    "person": "https://opus-project.eu/person/",
    "place":  "https://opus-project.eu/place/",
    "inst":   "https://opus-project.eu/institution/",
    "agg":    "https://opus-project.eu/aggregation/",
    "kw":     "https://opus-project.eu/keyword/",
}

CERTAINTY = {
    "certain":     (None, None, None),
    "inferred":    (1, None, None),
    "uncertain":   (None, 1, None),
    "approximate": (None, None, 1),
}


# ── tiny utils ────────────────────────────────────────────────────────────────

def val(x):
    s = str(x).strip() if x is not None else ""
    return None if s in ("", "nan", "NaN", "None") else s

def split(x, sep="|"):
    v = val(x)
    return [p.strip() for p in v.split(sep) if p.strip()] if v else []

def pipe2semi(x):
    parts = split(x)
    return ";".join(parts) if parts else None

def uri(kind, local):
    return URIRef(BASE[kind] + re.sub(r"[^\w\-.]", "_", str(local)))

def clean_url(x):
    m = re.search(r"https?://[^\s\"'<>)]+", str(x or ""))
    return URIRef(m.group(0).rstrip(".,;")) if m else None

def parse_date(x):
    v = val(x)
    if not v:
        return None, None, None
    m = re.fullmatch(r"(\d{4})(?:-(\d{2})(?:-(\d{2}))?)?", v)
    if not m:
        return None, None, None
    y, mo, d = m.group(1), m.group(2), m.group(3)
    if mo and not (1 <= int(mo) <= 12): return y, None, None
    if d  and not (1 <= int(d)  <= 31): return y, mo, None
    return y, mo, d

def certainty_flags(x):
    return CERTAINTY.get((val(x) or "").lower(), (None, None, None))

def is_letter_row(row):
    v = val(row.get("letter_manifestation_ID"))
    return bool(v) and not v.endswith(":")


# ── load ──────────────────────────────────────────────────────────────────────

def load(path):
    sheets = pd.read_excel(path, sheet_name=None, dtype=str)
    letters = sheets["Letter manifestations"]
    letters = letters[letters.apply(is_letter_row, axis=1)].reset_index(drop=True)
    return {
        "letters":      letters,
        "people":       sheets["People"],
        "places":       sheets["Places"],
        "institutions": sheets["Institutions"],
        "aggregations": sheets["Letter aggregations"],
        "people_forms": sheets["People Forms"],
        "places_forms": sheets["Places Forms"],
    }


# ── RDF ───────────────────────────────────────────────────────────────────────

def build_rdf(d):
    g = Graph()
    for prefix, ns in [("opus", OPUS), ("schema", SCHEMA), ("dcterms", DCTERMS),
                       ("skos", SKOS), ("bibo", BIBO), ("geo", GEO), ("owl", OWL)]:
        g.bind(prefix, ns)

    # places
    for _, r in d["places"].iterrows():
        if not val(r.get("place_ID")): continue
        u = uri("place", r["place_ID"])
        g.add((u, RDF.type, SCHEMA.Place))
        if val(r.get("place_name")):
            g.add((u, RDFS.label, Literal(r["place_name"])))
        if url := clean_url(r.get("wikidata_url")):
            g.add((u, OWL.sameAs, url))
        if val(r.get("latitude")):
            g.add((u, GEO.lat,  Literal(r["latitude"],  datatype=XSD.decimal)))
        if val(r.get("longitude")):
            g.add((u, GEO.long, Literal(r["longitude"], datatype=XSD.decimal)))
        if val(r.get("notes")):
            g.add((u, SKOS.note, Literal(r["notes"])))

    # people
    for _, r in d["people"].iterrows():
        if not val(r.get("person_ID")): continue
        u = uri("person", r["person_ID"])
        g.add((u, RDF.type, SCHEMA.Person))
        if val(r.get("person_name")):
            g.add((u, RDFS.label, Literal(r["person_name"])))
        if url := clean_url(r.get("wikidata_url")):
            g.add((u, OWL.sameAs, url))
        for field, pred in [("birthdate", SCHEMA.birthDate), ("deathdate", SCHEMA.deathDate)]:
            if val(r.get(field)) and re.fullmatch(r"\d{4}", val(r.get(field))):
                g.add((u, pred, Literal(val(r[field]), datatype=XSD.gYear)))
        for field, pred in [("birthplace_id", SCHEMA.birthPlace), ("deathplace_id", SCHEMA.deathPlace)]:
            if val(r.get(field)):
                g.add((u, pred, uri("place", r[field])))
        if val(r.get("role/title")):
            g.add((u, SCHEMA.jobTitle, Literal(r["role/title"])))
        if val(r.get("notes")):
            g.add((u, SKOS.note, Literal(r["notes"])))

    # institutions
    for _, r in d["institutions"].iterrows():
        if not val(r.get("institution_ID")): continue
        u = uri("inst", r["institution_ID"])
        g.add((u, RDF.type, SCHEMA.Organization))
        if val(r.get("institution_name")):
            g.add((u, SCHEMA.name, Literal(r["institution_name"])))
        if url := clean_url(r.get("wikidata_url")):
            g.add((u, OWL.sameAs, url))
        if val(r.get("location_ID")):
            g.add((u, SCHEMA.location, uri("place", r["location_ID"])))

    # people/places forms indexed by letter
    pf = d["people_forms"].groupby("letter_manifestation_ID")
    plf = d["places_forms"].groupby("letter_manifestation_ID")

    # letters
    for _, r in d["letters"].iterrows():
        lid = val(r.get("letter_manifestation_ID"))
        u = uri("letter", lid)
        g.add((u, RDF.type, OPUS.LetterManifestation))
        g.add((u, RDF.type, SCHEMA.CreativeWork))

        # date
        y, mo, d_ = parse_date(r.get("date_of_letter"))
        if y:
            date_str = y + (f"-{mo}" if mo else "") + (f"-{d_}" if d_ else "")
            dtype = XSD.date if (y and mo and d_) else XSD.gYear
            g.add((u, SCHEMA.dateCreated, Literal(date_str, datatype=dtype)))
        if val(r.get("date_as_marked")):
            g.add((u, OPUS.dateAsMarked, Literal(r["date_as_marked"])))
        inf, unc, approx = certainty_flags(r.get("date_of_letter_certainty"))
        if inf:   g.add((u, OPUS.dateIsInferred,    Literal(True, datatype=XSD.boolean)))
        if unc:   g.add((u, OPUS.dateIsUncertain,   Literal(True, datatype=XSD.boolean)))
        if approx:g.add((u, OPUS.dateIsApproximate, Literal(True, datatype=XSD.boolean)))

        # author / recipient / origin / destination — blank node for certainty
        for pid_field, cert_field, marked_field, pred, marked_pred in [
            ("author_ID",            "author_certainty",      "author_as_marked",      SCHEMA.author,          OPUS.authorAsMarked),
            ("recipient_ID",         "recipient_certainty",   "recipient_as_marked",   OPUS.recipient,         OPUS.recipientAsMarked),
        ]:
            if val(r.get(pid_field)):
                bn = BNode()
                g.add((u,  pred,        bn))
                g.add((bn, OPUS.value,  uri("person", r[pid_field])))
                cert = val(r.get(cert_field))
                if cert:
                    g.add((bn, OPUS.certainty, Literal(cert)))
            if val(r.get(marked_field)):
                g.add((u, marked_pred, Literal(r[marked_field])))

        for pid_field, cert_field, pred, marked_pred in [
            ("origin_place_ID",      "origin_place_certainty",      SCHEMA.locationCreated, OPUS.originAsMarked),
            ("destination_place_ID", "destination_place_certainty", OPUS.destination,       OPUS.destinationAsMarked),
        ]:
            if val(r.get(pid_field)):
                bn = BNode()
                g.add((u,  pred,        bn))
                g.add((bn, OPUS.value,  uri("place", r[pid_field])))
                cert = val(r.get(cert_field))
                if cert:
                    g.add((bn, OPUS.certainty, Literal(cert)))
        if val(r.get("origin_place_name")):
            g.add((u, OPUS.originAsMarked,      Literal(r["origin_place_name"])))
        if val(r.get("destination_place_name")):
            g.add((u, OPUS.destinationAsMarked, Literal(r["destination_place_name"])))

        # simple literals
        for field, pred in [
            ("abstract",        DCTERMS.abstract),
            ("incipit",         OPUS.incipit),
            ("notes",           SKOS.note),
            ("transcription",   OPUS.transcription),
            ("print_information", BIBO.edition),
            ("shelfmark",       BIBO.locator),
        ]:
            if val(r.get(field)):
                g.add((u, pred, Literal(r[field])))

        for lang in split(r.get("languages")):
            g.add((u, SCHEMA.inLanguage, Literal(lang)))

        # keywords → SKOS concepts
        kws = set(split(r.get("keywords_manual")) + split(r.get("keywords_aggregated")))
        for kw in kws:
            ku = uri("kw", kw.lower().replace(" ", "_"))
            g.add((ku, RDF.type,       SKOS.Concept))
            g.add((ku, SKOS.prefLabel, Literal(kw, lang="en")))
            g.add((u,  DCTERMS.subject, ku))

        for pid in split(r.get("people_mentioned_IDs")):
            g.add((u, SCHEMA.mentions, uri("person", pid)))
        for plid in split(r.get("places_mentioned_IDs")):
            g.add((u, SCHEMA.mentions, uri("place", plid)))

        if val(r.get("repository_ID")):
            g.add((u, DCTERMS.isPartOf, uri("inst", r["repository_ID"])))
        if url := clean_url(r.get("letter_url")):
            g.add((u, SCHEMA.url, url))

        # people forms — name as used in this letter (blank node)
        if lid in pf.groups:
            for _, pf_row in pf.get_group(lid).iterrows():
                if val(pf_row.get("person_ID")) and val(pf_row.get("person_name_form")):
                    bn = BNode()
                    g.add((u,  OPUS.hasNameAppearance, bn))
                    g.add((bn, OPUS.person,            uri("person", pf_row["person_ID"])))
                    g.add((bn, OPUS.nameAsUsed,        Literal(pf_row["person_name_form"])))

        # places forms
        if lid in plf.groups:
            for _, plf_row in plf.get_group(lid).iterrows():
                if val(plf_row.get("place_ID")) and val(plf_row.get("place_name_form")):
                    bn = BNode()
                    g.add((u,  OPUS.hasPlaceAppearance, bn))
                    g.add((bn, OPUS.place,              uri("place", plf_row["place_ID"])))
                    g.add((bn, OPUS.placeNameAsUsed,    Literal(plf_row["place_name_form"])))

    # aggregations
    for agg_id, grp in d["aggregations"].groupby("letter_aggregation_id"):
        if not val(agg_id): continue
        au = uri("agg", agg_id)
        g.add((au, RDF.type, OPUS.LetterAggregation))
        for _, r in grp.iterrows():
            if val(r.get("letter_manifestation_ID")):
                g.add((au, OPUS.hasManifestation, uri("letter", r["letter_manifestation_ID"])))

    return g


# ── EMLO ─────────────────────────────────────────────────────────────────────

def build_emlo(d):
    letters_out = []
    for _, r in d["letters"].iterrows():
        y, mo, day = parse_date(r.get("date_of_letter"))
        d_inf, d_unc, d_app = certainty_flags(r.get("date_of_letter_certainty"))
        a_inf, a_unc, _     = certainty_flags(r.get("author_certainty"))
        rc_inf, rc_unc, _   = certainty_flags(r.get("recipient_certainty"))
        o_inf, o_unc, _     = certainty_flags(r.get("origin_place_certainty"))
        t_inf, t_unc, _     = certainty_flags(r.get("destination_place_certainty"))
        is_print = val(r.get("print")) == "print"

        letters_out.append({
            "LETTER Number":                   val(r.get("letter_manifestation_ID")),
            "LETTER ID from PARTNER PROJECT":  val(r.get("letter_in_collection")),
            "DATE AS MARKED ON LETTER":        val(r.get("date_as_marked")),
            "CALENDAR":                        "G",
            "YEAR DATE OF LETTER":             y,
            "MONTH DATE OF LETTER":            mo,
            "DAY DATE OF LETTER":              day,
            "DATE INFERRED":                   d_inf,
            "DATE UNCERTAIN":                  d_unc,
            "DATE APPROXIMATE":                d_app,
            "AUTHOR":                          val(r.get("author_name")),
            "EMLO ID of AUTHOR":               None,
            "AUTHOR AS MARKED":                val(r.get("author_as_marked")),
            "AUTHOR INFERRED":                 a_inf,
            "AUTHOR UNCERTAIN":                a_unc,
            "RECIPIENT":                       val(r.get("recipient_name")),
            "EMLO ID of RECIPIENT":            None,
            "RECIPIENT AS MARKED":             val(r.get("recipient_as_marked")),
            "RECIPIENT INFERRED":              rc_inf,
            "RECIPIENT UNCERTAIN":             rc_unc,
            "ORIGIN NAME":                     val(r.get("origin_place_name")),
            "EMLO ID of ORIGIN PLACE":         None,
            "ORIGIN AS MARKED":                val(r.get("origin_place_name")),
            "ORIGIN INFERRED":                 o_inf,
            "ORIGIN UNCERTAIN":                o_unc,
            "DESTINATION NAME":                val(r.get("destination_place_name")),
            "EMLO ID of DESTINATION PLACE":    None,
            "DESTINATION AS MARKED":           val(r.get("destination_place_name")),
            "DESTINATION INFERRED":            t_inf,
            "DESTINATION UNCERTAIN":           t_unc,
            "ABSTRACT":                        val(r.get("abstract")),
            "KEYWORDS":                        pipe2semi(r.get("keywords_aggregated")),
            "LANGUAGES":                       pipe2semi(r.get("languages")),
            "INCIPIT":                         val(r.get("incipit")),
            "NOTES ON LETTER":                 val(r.get("notes")),
            "PEOPLE MENTIONED":                pipe2semi(r.get("people_mentioned_names")),
            "EMLO ID PERSON MENTIONED":        None,
            "RELATED RESOURCE URL":            val(r.get("letter_url")),
            "MS MANIFESTATION":                None if is_print else "ALS",
            "REPOSITORY":                      val(r.get("repository_name")),
            "SHELFMARK":                       val(r.get("shelfmark")),
            "PRINT or DIGITAL MANIFESTATION":  "P" if is_print else None,
            "PRINTED or DIGITAL COPY DETAILS": val(r.get("print_information")),
        })

    places_lookup = dict(zip(d["places"]["place_ID"], d["places"]["place_name"]))

    people_out = []
    for _, r in d["people"].iterrows():
        if not val(r.get("person_ID")): continue
        name = val(r.get("person_name"))
        primary = name
        if name and "," not in name:
            parts = name.rsplit(" ", 1)
            if len(parts) == 2:
                primary = f"{parts[1]}, {parts[0]}"
        people_out.append({
            "NAME as appears in LETTER TAB":     name,
            "EMLO PERSON ID":                    None,
            "PRIMARY NAME":                      primary,
            "BIRTH YEAR":                        val(r.get("birthdate")),
            "DEATH YEAR":                        val(r.get("deathdate")),
            "OCCUPATIONS, ROLES, and/or TITLES": val(r.get("role/title")),
            "GENERAL NOTES ON PERSON":           val(r.get("notes")),
            "RELATED RESOURCE URL":              val(r.get("wikidata_url")),
            "Author ID from partner project":    val(r.get("person_ID")),
        })

    places_out = []
    for _, r in d["places"].iterrows():
        if not val(r.get("place_ID")): continue
        notes = val(r.get("notes")) or ""
        geonames = None
        if "geonames.org" in notes:
            m = re.search(r"https?://[^\s]+geonames[^\s)]+", notes)
            if m: geonames = m.group(0).rstrip(".,;)")
        places_out.append({
            "NAME OF PLACE as it appears in LETTERS TAB": val(r.get("place_name")),
            "EMLO place ID":                              None,
            "TOWN, CITY, HAMLET, OR VILLAGE":            val(r.get("place_name")),
            "LATITUDE":                                   val(r.get("latitude")),
            "LONGITUDE":                                  val(r.get("longitude")),
            "GENERAL NOTES ON PLACE":                     notes or None,
            "RELATED RESOURCE NAME":                      "GeoNames" if geonames else ("Wikidata" if val(r.get("wikidata_url")) else None),
            "RELATED RESOURCE URL":                       geonames or val(r.get("wikidata_url")),
            "PLACE ID FROM PARTNER PROJECT":              val(r.get("place_ID")),
        })

    repos_out = []
    for _, r in d["institutions"].iterrows():
        if not val(r.get("institution_ID")): continue
        city = places_lookup.get(val(r.get("location_ID")))
        repos_out.append({
            "REPOSITORY NAME":            val(r.get("institution_name")),
            "EMLO place ID":              None,
            "PRIMARY NAME OF REPOSITORY": val(r.get("institution_name")),
            "CITY OR TOWN OF REPOSITORY": city,
            "RELATED RESOURCE URL":       val(r.get("wikidata_url")),
        })

    return {
        "Letters":      pd.DataFrame(letters_out),
        "People":       pd.DataFrame(people_out),
        "Places":       pd.DataFrame(places_out),
        "Repositories": pd.DataFrame(repos_out),
    }


def write_emlo(tables, path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, df in tables.items():
        ws = wb.create_sheet(sheet_name)
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        for ci, col in enumerate(df.columns, 1):
            c = ws.cell(row=1, column=ci, value=col)
            c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=9)
            c.fill      = PatternFill("solid", fgColor="2C3E50")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(ci)].width = max(14, min(len(col), 28))
        ws.row_dimensions[1].height = 28
        for ri, (_, row) in enumerate(df.iterrows(), 2):
            bg = "ECF0F1" if ri % 2 == 0 else "FFFFFF"
            for ci, v in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=None if pd.isna(v) or v is None else v)
                c.font      = Font(name="Arial", size=9)
                c.fill      = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(path)


# ── main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else "OPUS_data_model.xlsx"
    emlo_out = sys.argv[2] if len(sys.argv) > 2 else "output_emlo.xlsx"
    rdf_out  = sys.argv[3] if len(sys.argv) > 3 else "output_graph.ttl"

    print(f"Loading {path} …")
    d = load(path)

    print("Building RDF …")
    g = build_rdf(d)
    g.serialize(destination=rdf_out, format="turtle")

    print("Building EMLO …")
    tables = build_emlo(d)
    write_emlo(tables, emlo_out)

    print(f"\nDone.")
    print(f"  Letters : {len(tables['Letters'])}")
    print(f"  People  : {len(tables['People'])}")
    print(f"  Places  : {len(tables['Places'])}")
    print(f"  Triples : {len(g)}")
    print(f"  → {emlo_out}")
    print(f"  → {rdf_out}")
